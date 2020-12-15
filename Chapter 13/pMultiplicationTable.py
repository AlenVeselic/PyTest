#! python3

# pMultiplicationTable.py N - creates a multiplication table for N and outputs an excel sheet of it


import openpyxl, sys

from openpyxl.utils import get_column_letter, column_index_from_string



multiplyNumber = int(sys.argv[1])

workbook = openpyxl.Workbook()

sheet = workbook['Sheet']


for row in range(1, multiplyNumber + 2):
    for column in range(1,multiplyNumber + 2):

        cell = sheet["" + get_column_letter(column) + str(row)]
        if row == 1 and column == 1:
            continue
        if row == 1:
            cell.value = column - 1
            continue
        if column == 1:
            cell.value = row - 1
            continue

        cell.value = '=' + get_column_letter(column) + str(1) + '*' + get_column_letter(1) + str(row)

workbook.save('multiPly.xlsx')
        

