# sheet inverter - inverts spreadsheet cell values

import openpyxl

from openpyxl.utils import get_column_letter

oldWorkbook = openpyxl.load_workbook('Chapter 13/sellings.xlsx')
oldSheet = oldWorkbook.active

invertedWorkbook = openpyxl.Workbook()
inSheet = invertedWorkbook.active

for row in range(1, oldSheet.max_row + 1):
    for column in range(1, oldSheet.max_column + 1):
        newCell = inSheet[get_column_letter(row) + str(column)]
        oldCell = oldSheet[get_column_letter(column) + str(row)]

        newCell.value = oldCell.value

invertedWorkbook.save('inverted.xlsx')