import openpyxl

from openpyxl.utils import get_column_letter

from pathlib import Path

workbook = openpyxl.Workbook()

sheet = workbook.active

fileNumber = 3

filePaths = ( Path(r"H:\The Folder of many text files\Conspiracy turned scientific fact.txt"),
              Path(r"H:\The Folder of many text files\Neighbor report ticket #0059.txt"),
              Path(r"H:\The Folder of many text files\Order.txt")

)

for f in range(fileNumber):
    textFile = open(filePaths[f])
    lines = textFile.readlines()

    for line in range(len(lines)):
        cell = sheet[get_column_letter(f + 1) + str(line + 1)]
        cell.value = lines[line]

workbook.save('text2sheetTest.xlsx')


