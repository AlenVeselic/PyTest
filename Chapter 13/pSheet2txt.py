import openpyxl

from openpyxl.utils import get_column_letter

workbook = openpyxl.load_workbook("text2sheetTest.xlsx")
sheet = workbook.active


for column in range(1, sheet.max_column + 1):
    outputString = ""
    print('Working on the ' + str(column) + '. file...')
    for row in range(1, sheet.max_row + 1):
        cell = sheet[get_column_letter(column) + str(row)]
        if cell.value != None:
            outputString += cell.value
        else:
            outputString += "\n"
    print('Finishing the ' + str(column) + '. file...')
    fileAssembly = open("sheetText" + str(column) + '.txt', 'w')
    fileAssembly.write(outputString)

print('Done.')