#! python3

# pBlankRows.py rowI rowNum Sheet - inserts a a number of blank rows into a given index into a given sheet

import sys, openpyxl, logging, re

from openpyxl.utils import get_column_letter 

logging.basicConfig(level = logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')


blankRowIndex = int(sys.argv[1])
blankRowNumber = int(sys.argv[2])

naming = re.compile(r'^\w+\.')

sheetName = sys.argv[3]

filename = naming.search(sheetName)

workbook = openpyxl.load_workbook(sheetName)

oldSheet = workbook.active

newWorkbook = openpyxl.Workbook()

newSheet = newWorkbook.active

for row in range(1,oldSheet.max_row + blankRowNumber + 1):
    logging.debug('Copying row number: ' + str(row))
    for column in range(1,oldSheet.max_column + 1):
        if row < blankRowIndex:
            newCell = newSheet[get_column_letter(column) + str(row)]
        else:
            newCell = newSheet[get_column_letter(column) + str(row + blankRowNumber)]
        oldCell = oldSheet[get_column_letter(column) + str(row)]
        newCell.value = oldCell.value

print('Done.')
if filename != None:
    length = len(filename.group())
    newWorkbook.save(filename.group()[:length-1] + '-cpy.xlsx')
else:
    newWorkbook.save('rowsInserted.xlsx')