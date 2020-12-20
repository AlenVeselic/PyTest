#! python3

# pExcel2CSV.py - converts all .xlsx files in .csv files

import openpyxl, os, csv

os.makedirs('convertedCSVs', exist_ok=True)

for fileName in os.listdir('.'):
    if not fileName.endswith(".xlsx"):
        continue
    else:
        wb = openpyxl.load_workbook(fileName)
        sheetNum = len(wb.sheetnames)

        print("Converting file " + fileName + '...')

        for sheet in range(sheetNum):
            print("Converting file's " + wb.sheetnames[sheet] + " sheet ...")
            currentSheet = wb[wb.sheetnames[sheet]]
            csvFilename = fileName[:-5] + '_' + wb.sheetnames[sheet] + ".csv"
            csvFile = open(os.path.join('convertedCSVs',csvFilename), 'w', newline='')
            csvWriter = csv.writer(csvFile)
            for row in range(1, currentSheet.max_row):
                rows = []
                for column in range(1, currentSheet.max_column):
                    rows.append(currentSheet.cell(column = column, row =row).value)
                csvWriter.writerow(rows)
                    
            csvFile.close()

            

